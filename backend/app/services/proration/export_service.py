"""Export service for generating CSV, Excel, and PDF files."""

from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from app.models.proration import MineralHolderRow

HEADERS = [
    "Owner",
    "Year",
    "Appraisal Value",
    "Interest Type",
    "County",
    "Block",
    "Section",
    "Abstract",
    "Legal Description",
    "Property",
    "Operator",
    "Raw RRC",
    "New Record",
    "Estimated Monthly Revenue",
    "Interest",
    "RRC Acres",
    "Est NRA",
    "Notes",
]

# Column indices (1-based) for SUBTOTAL formulas
_COL_EST_MONTHLY_REV = 14  # "Estimated Monthly Revenue" is column N
_COL_EST_NRA = 17  # "Est NRA" is column Q


def _row_values(row: MineralHolderRow) -> list:
    """Extract export values from a MineralHolderRow."""
    return [
        row.owner,
        row.year,
        row.appraisal_value,
        row.interest_type,
        row.county,
        row.block,
        row.section,
        row.abstract,
        row.legal_description,
        row.property,
        row.operator,
        row.raw_rrc,
        row.new_record,
        row.estimated_monthly_revenue,
        row.interest,
        row.rrc_acres,
        row.est_nra,
        row.notes,
    ]


def to_csv(rows: list[MineralHolderRow]) -> bytes:
    """Export mineral holder rows to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(_row_values(row))
    return output.getvalue().encode("utf-8")


def to_excel(rows: list[MineralHolderRow], sheet_name: str = "MH") -> bytes:
    """Export mineral holder rows to Excel format.

    Matches the standard Mineral Holder export layout:
    - Column order: Owner, Year, Appraisal Value, Interest Type, County,
      Block, Section, Abstract, Legal Description, Property, Operator,
      Raw RRC, New Record, Estimated Monthly Revenue, Interest, RRC Acres,
      Est NRA, Notes
    - SUBTOTAL formulas for Estimated Monthly Revenue and Est NRA
    """
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit

    ws.append(HEADERS)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for row in rows:
        ws.append(_row_values(row))

    # Add SUBTOTAL row (SUBTOTAL function 9 = SUM ignoring hidden rows)
    if rows:
        data_last_row = len(rows) + 1  # +1 for header row
        subtotal_row_idx = data_last_row + 1

        rev_col = get_column_letter(_COL_EST_MONTHLY_REV)
        nra_col = get_column_letter(_COL_EST_NRA)

        # Estimated Monthly Revenue subtotal
        rev_cell = ws.cell(
            row=subtotal_row_idx, column=_COL_EST_MONTHLY_REV,
            value=f"=SUBTOTAL(9,{rev_col}2:{rev_col}{data_last_row})",
        )
        rev_cell.font = Font(bold=True)

        # Est NRA subtotal
        nra_cell = ws.cell(
            row=subtotal_row_idx, column=_COL_EST_NRA,
            value=f"=SUBTOTAL(9,{nra_col}2:{nra_col}{data_last_row})",
        )
        nra_cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def to_pdf(rows: list[MineralHolderRow]) -> bytes:
    """Export mineral holder rows to PDF format."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    data = [
        [
            "Owner",
            "County",
            "Interest",
            "RRC Acres",
            "Est NRA",
            "$/NRA",
            "Appraisal Value",
        ]
    ]

    for row in rows:
        data.append(
            [
                row.owner or "",
                row.county or "",
                f"{row.interest:.4f}" if row.interest else "",
                f"{row.rrc_acres:.2f}" if row.rrc_acres else "",
                f"{row.est_nra:.4f}" if row.est_nra else "",
                f"${row.dollars_per_nra:.2f}" if row.dollars_per_nra else "",
                f"${row.appraisal_value:.2f}" if row.appraisal_value else "",
            ]
        )

    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return buffer.read()
